#!/usr/bin/env python3
"""
parse_form.py — Parse a SurveyCTO/ODK form definition (.xlsx) into
readable CSV files and a plain-text summary.

Usage:
    python3 parse_form.py <form.xlsx> [--outdir /tmp/form_parsed]

Outputs (in outdir):
    survey.csv   — the survey sheet
    choices.csv  — the choices sheet
    settings.csv — the settings sheet
    summary.txt  — human-readable overview of form structure
"""

import argparse
import csv
import os
import sys
from collections import Counter, OrderedDict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


def read_sheet(wb, sheet_name):
    """Read a worksheet into a list of dicts. Returns (headers, rows)."""
    # Case-insensitive sheet lookup
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    actual_name = sheet_map.get(sheet_name.lower())
    if actual_name is None:
        return [], []
    ws = wb[actual_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            d[h] = "" if val is None else str(val).strip()
        # Skip completely blank rows
        if any(v for v in d.values()):
            data.append(d)
    return headers, data


def write_csv(headers, rows, path):
    """Write list of dicts to CSV."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def classify_type(raw_type):
    """Return a simplified category for a field type string."""
    t = raw_type.lower().strip()
    if t.startswith("select_one"):
        return "select_one"
    if t.startswith("select_multiple"):
        return "select_multiple"
    if t in ("begin group", "begin_group"):
        return "begin_group"
    if t in ("end group", "end_group"):
        return "end_group"
    if t in ("begin repeat", "begin_repeat"):
        return "begin_repeat"
    if t in ("end repeat", "end_repeat"):
        return "end_repeat"
    if t in ("text", "integer", "decimal", "date", "datetime", "time",
             "geopoint", "geotrace", "geoshape", "image", "audio",
             "video", "barcode", "calculate", "calculate_here", "note",
             "acknowledge", "rank", "file", "hidden", "xml-external"):
        return t
    # Metadata types
    if t in ("start", "end", "today", "deviceid", "subscriberid",
             "simserial", "phonenumber", "username", "caseid",
             "audit", "speed violations count", "speed violations list",
             "speed violations audit", "text audit"):
        return "metadata"
    return "other"


def get_list_name(raw_type):
    """Extract the choice list name from a select_one/select_multiple type."""
    parts = raw_type.strip().split()
    if len(parts) >= 2 and parts[0] in ("select_one", "select_multiple"):
        return parts[1]
    return None


def get_label(row, headers):
    """Get the best available label from a row (handles multilingual forms)."""
    # Try 'label' first, then any 'label::*' column
    if row.get("label"):
        return row["label"]
    for h in headers:
        if h.lower().startswith("label") and row.get(h):
            return row[h]
    return ""


def generate_summary(survey_headers, survey_rows, choices_headers,
                     choices_rows, settings_headers, settings_rows):
    """Generate a human-readable summary of the form."""
    lines = []
    lines.append("=" * 60)
    lines.append("SURVEYCTO FORM DEFINITION — SUMMARY")
    lines.append("=" * 60)

    # Settings
    if settings_rows:
        s = settings_rows[0]
        title = s.get("form_title", s.get("title", ""))
        form_id = s.get("form_id", s.get("id", ""))
        version = s.get("version", "")
        lang = s.get("default_language", "")
        lines.append(f"\nForm title:    {title}")
        lines.append(f"Form ID:       {form_id}")
        if version:
            lines.append(f"Version:       {version}")
        if lang:
            lines.append(f"Default lang:  {lang}")

    # Count fields by type
    type_counts = Counter()
    visible_fields = 0
    metadata_fields = 0
    groups = []
    repeats = []
    group_stack = []

    for row in survey_rows:
        raw_type = row.get("type", "")
        cat = classify_type(raw_type)
        type_counts[cat] += 1
        name = row.get("name", "")
        label = get_label(row, survey_headers)

        if cat == "begin_group":
            group_stack.append(name)
            groups.append({
                "name": name,
                "label": label,
                "depth": len(group_stack),
                "fields": 0,
                "relevance": row.get("relevance", "")
            })
        elif cat == "end_group":
            if group_stack:
                group_stack.pop()
        elif cat == "begin_repeat":
            repeats.append({
                "name": name,
                "label": label,
                "relevance": row.get("relevance", "")
            })
            group_stack.append(name)
        elif cat == "end_repeat":
            if group_stack:
                group_stack.pop()
        elif cat == "metadata":
            metadata_fields += 1
        elif cat not in ("end_group", "end_repeat", "other"):
            visible_fields += 1
            # Attribute to innermost group
            if groups and group_stack:
                for g in reversed(groups):
                    if g["name"] == group_stack[-1]:
                        g["fields"] += 1
                        break

    lines.append(f"\n--- Field counts ---")
    lines.append(f"Total fields (visible):  {visible_fields}")
    lines.append(f"Metadata fields:         {metadata_fields}")
    lines.append(f"Groups:                  {type_counts.get('begin_group', 0)}")
    lines.append(f"Repeat groups:           {type_counts.get('begin_repeat', 0)}")
    lines.append(f"\nBy type:")
    for t in sorted(type_counts.keys()):
        if t not in ("begin_group", "end_group", "begin_repeat",
                      "end_repeat", "metadata"):
            lines.append(f"  {t:25s} {type_counts[t]}")

    # Top-level groups (sections)
    top_groups = [g for g in groups if g["depth"] == 1]
    if top_groups:
        lines.append(f"\n--- Sections (top-level groups) ---")
        for i, g in enumerate(top_groups, 1):
            label_str = g["label"] or g["name"]
            lines.append(f"  {i}. {label_str} [{g['name']}] "
                         f"— {g['fields']} fields")
            if g["relevance"]:
                lines.append(f"     Skip: {g['relevance']}")

    # Repeat groups
    if repeats:
        lines.append(f"\n--- Repeat groups ---")
        for r in repeats:
            label_str = r["label"] or r["name"]
            lines.append(f"  • {label_str} [{r['name']}]")

    # Choice lists summary
    if choices_rows:
        list_names = Counter(r.get("list_name", "") for r in choices_rows)
        lines.append(f"\n--- Choice lists ---")
        lines.append(f"Total lists: {len(list_names)}")
        for ln, count in sorted(list_names.items()):
            lines.append(f"  {ln:30s} {count} options")

    # Variable listing (non-metadata, non-group)
    lines.append(f"\n--- Variable listing ---")
    lines.append(f"{'name':30s} {'type':25s} label")
    lines.append("-" * 100)
    for row in survey_rows:
        raw_type = row.get("type", "")
        cat = classify_type(raw_type)
        if cat in ("begin_group", "end_group", "begin_repeat",
                    "end_repeat", "metadata"):
            continue
        name = row.get("name", "")
        label = get_label(row, survey_headers)
        # Truncate label for summary
        if len(label) > 60:
            label = label[:57] + "..."
        lines.append(f"  {name:30s} {raw_type:25s} {label}")

    lines.append(f"\n{'=' * 60}")
    lines.append("END OF SUMMARY")
    lines.append(f"{'=' * 60}\n")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Parse a SurveyCTO/ODK .xlsx form definition")
    parser.add_argument("xlsx", help="Path to the form definition .xlsx file")
    parser.add_argument("--outdir", default="/tmp/form_parsed",
                        help="Output directory for CSVs and summary")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Error: file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.outdir, exist_ok=True)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    # Parse each sheet
    survey_h, survey_r = read_sheet(wb, "survey")
    choices_h, choices_r = read_sheet(wb, "choices")
    settings_h, settings_r = read_sheet(wb, "settings")

    wb.close()

    # Write CSVs
    if survey_h:
        write_csv(survey_h, survey_r, os.path.join(args.outdir, "survey.csv"))
        print(f"  survey.csv   — {len(survey_r)} rows")
    else:
        print("  WARNING: no 'survey' sheet found!")

    if choices_h:
        write_csv(choices_h, choices_r, os.path.join(args.outdir, "choices.csv"))
        print(f"  choices.csv  — {len(choices_r)} rows")
    else:
        print("  WARNING: no 'choices' sheet found!")

    if settings_h:
        write_csv(settings_h, settings_r,
                  os.path.join(args.outdir, "settings.csv"))
        print(f"  settings.csv — {len(settings_r)} rows")

    # Generate summary
    summary = generate_summary(survey_h, survey_r, choices_h, choices_r,
                                settings_h, settings_r)
    summary_path = os.path.join(args.outdir, "summary.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary)
    print(f"  summary.txt  — form overview")

    print(f"\nAll files written to: {args.outdir}")


if __name__ == "__main__":
    main()
